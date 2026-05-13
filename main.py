import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
from collections import defaultdict
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.formula import ArrayFormula
from copy import copy


NAME_MAP = {
    "Solapur I": "SOLAPUR",
    "National Capital Thermal Power Station Dadri II (NR)": "DADRT2",
    "Unchahar III (NR)": "UNCHAHAR3",
    "Indira Gandhi STPS (Jhajjar) (NR)": "JHAJJAR",
    "BLA": "BLA_POWER_I&II",
    "BLA ": "BLA_POWER_I&II",
    "Unchahar II (NR)": "UNCHAHAR2",
    "DVC (DTPS) (STOA LOI)": "DVC",
    "Ghatampur Thermal Power Station (NUPPL)": "NUPPL_UP_1",
    "Unchahar I (NR)": "UNCHAHAR1",
    "Khargone": "KHARGONE",
    "Unchahar IV (NR)": "UNCHAHAR4",
    "Tanda II": "TANDA2",
    "Gadarwara": "GADARWARA",
    "JP Bina": "JP_BINA_TPP",
    "TRN Energy Pvt. Ltd. (Medium Term)": "TRN_ENERGY",
    "SSTPS Khandwa (Singaji)": "SSTPP-I",
    "MB Power": "MB POWER",
    "Jhabua Power": "JHABUA_IPP",
    "SSTPS Khandwa (Singaji) # 3 & 4": "SSTPP-II",
    "Mauda-I": "MOUDA1",
    "Mauda-II": "MOUDA2",
    "SGTPS -4x210": "SGT",
    "STPS Sarni 10 & 11": "STP-IV",
    "KAHALGAON-II": "KHSTPP_II",
    "SKS Power Generation (Medium Term)": "SKS_Raigarh",
    "RKM Powergen (Talcher Field) (Medium Term)": "RKM_POWER_T",
    "RKM Powergen  (IB Field) (Medium Term)": "RKM_POWER_IB",
    "Jindal Power Limited (Medium Term)": "JIPL",
    "Meja (NR)": "MEJA",
    "VSTPS-I": "VSTPS1",
    "Jindal India Thermal Power (Medium Term)": "JPL2",
    "VSTPS-V": "VSTPS5",
    "VSTPS-III": "VSTPS3",
    "VSTPS-II": "VSTPS2",
    "VSTPS-IV": "VSTPS4",
    "SGTPS -1x500": "SGT_II",
    "Khurja STPP": "KHURJA_STPP",
    "GANDHAR (APM)": "GANDHAR_APM",
    "Auraiya  APM (NR)": "AURY_GF",
    "Dadri APM (NR)": "DADRI_GF",
    "Anta APM (NR)": "ANTA_GF",
    "Singrauli (NR)": "SINGRAULI",
    "Rihand I (NR)": "RIHAND1",
    "Rihand II (NR)": "RIHAND2",
    "ATPS (210MW) Chachai": "AMK_III",
    "Rihand III (NR)": "RIHAND3",
    "KAWAS (APM)": "KAWAS_APM",
    "Sasan": "SASAN",
    "SIPAT II": "SIPAT2",
    "SIPAT I": "SIPAT1",
    "KSTPS": "KSTPS7",
    "KSTPS-III": "KSTPS12",
    "Lara": "LARA",
    "JP Nigrie": "JNSTPP",
    "Koteshwar (NR)": "KOTESHWR",
    "Kishanganga (NR)": "KISHANGANGA",
    "Dulhasti (NR)": "DULHASTI",
    "Koldam (NR)": "KOLDAM",
    "URI II (NR)": "URI2",
    "Sewa II (NR)": "SEWA2",
    "Tehri I (NR)": "TEHRI",
    "Chamera III (NR)": "CHAMERA3",
    "Rampur HEP (NR)": "RAMPUR",
    "Parbati III (NR)": "PARBATI3",
    "Dhauliganga (NR)": "DHAULIGNGA",
    "Chamera II (NR)": "CHAMERA2",
    "Nathpa-Jhakri (NR)": "NJPC",
    "SSP": "SSP",
    "Bargi HPS": "BARGI HPS",
    "KAWAS (NAPM)": "KAWAS_NAPM",
    "GANDHAR (NAPM)": "GANDHAR_NAPM",
    "KAWAS (Liquid)": "KAWAS_LF",
    "Dadri Liquid (NR)": "DADRI_LF",
    "Auraiya  Liquid (NR)": "AURY_LF",
    "Anta Liquid (NR)": "ANTA_LF",
    "Auraiya  LNG (NR)": "AURY_RF",
    "Anta LNG (NR)": "ANTA_RF",
    "Dadri LNG (NR)": "DADRI_RF",
    "Auraiya  CLNG (NR)": "AURY_CRF",
    "KAWAS (LNG)": "KAWAS_RLNG",
    "GANDHAR (LNG)": "GANDHAR_RLNG",
    "Anta CLNG (NR)": "ANTA_CRF",
    "Dadri  Comm LNG (NR)": "DADRI_CRF",
    "KAWAS (CLNG)": "KAWAS_CRF",
    "GANDHAR (CLNG)": "GANDHAR_CRF",
    "Torrent (Indigenous)": "TORRENT_GUVNL",
    "Gandhi Sagar HPS": "GANDHI-SAGAR HPS",
    "GANDHI-SAGAR HPS": "GANDHI-SAGAR HPS",
    "Pench HPS": "PENCH HPS",
    "Birsinghpur HPS": "BIRSINGHPUR HPS",
    "Rajghat HPS": "RAJGHAT HPS",
    "Jhinna HPS": "Jhinha BSR-4",
    "Madikhera HPS": "Madikheda",
    "R P sagar (Inter state project)": "R P sagar (Inter state project)",
    "Jawahar Sagar (Inter state project)": "Jawahar Sagar (Inter state project)",
    " TONS HPS BSR-1": "TONS HPS BSR-1",
    "TONS HPS BSR-1": "TONS HPS BSR-1",
}

ACTIVE_NAME_MAP = {re.sub(r"\s+", " ", k.strip()): v for k, v in NAME_MAP.items()}
RANKED_STATIONS = [
    "SSP", "JNSTPP", "NJPC", "CHAMERA2", "DHAULIGNGA", "LARA", "KSTPS12", "PARBATI3",
    "KSTPS7", "SIPAT1", "SIPAT2", "SASAN", "KAWAS_APM", "RIHAND3", "AMK_III", "RIHAND2",
    "RIHAND1", "SINGRAULI", "ANTA_GF", "DADRI_GF", "RAMPUR", "AURY_GF", "CHAMERA3",
    "GANDHAR_APM", "KHURJA_STPP", "TEHRI", "SEWA2", "URI2", "SGT_II", "KOLDAM", "VSTPS4",
    "DULHASTI", "VSTPS2", "VSTPS3", "VSTPS5", "KISHANGANGA", "JPL2", "VSTPS1", "MEJA",
    "JIPL", "RKM_POWER_IB", "RKM_POWER_T", "SKS_Raigarh", "KHSTPP_II", "STP-IV", "SGT",
    "MOUDA2", "MOUDA1", "SSTPP-II", "JHABUA_IPP", "MB POWER", "SSTPP-I", "KOTESHWR",
    "TRN_ENERGY", "JP_BINA_TPP", "GADARWARA", "TANDA2", "UNCHAHAR4", "KHARGONE",
    "UNCHAHAR1", "NUPPL_UP_1", "DVC", "UNCHAHAR2", "BLA_POWER_I&II", "JHAJJAR", "UNCHAHAR3",
    "DADRT2", "SOLAPUR", "TORRENT_GUVNL", "GANDHAR_CRF", "KAWAS_CRF", "DADRI_CRF",
    "ANTA_CRF", "GANDHAR_RLNG", "KAWAS_RLNG", "AURY_CRF", "DADRI_RF", "ANTA_RF", "AURY_RF",
    "ANTA_LF", "AURY_LF", "DADRI_LF", "KAWAS_LF", "GANDHAR_NAPM", "KAWAS_NAPM",
]
ALLOWED_STATIONS: set[str] = set()
HYDEL_CODES = {
    # Normalize to match name corrector codes (spacing/case insensitive)
    "gandhi-sagar hps",
    "pench hps",
    "bargi hps",
    "tons hps bsr-1",
    "birsinghpur hps",
    "rajghat hps",
    "devlond hps bsr-3",
    "silpara hps bsr-2",
    "bargi lbcph",
    "jhinha bsr-4",
    "madikheda",
    "isp",
    "osp",
    "jawahar sagar (inter state project)",
    "r p sagar (inter state project)",
}
SSGS_CODES = {
    "amk_iii",
    "sgt",
    "sgt_ii",
    "stp-iv",
    "sstpp-i",
    "sstpp-ii",
}


def norm_key(text: object) -> str:
    return re.sub(r"\s+", " ", str(text or "").strip().lower())

# ENTT-2 column order and mode mapping (derived from reference workbook)
ENTT2_ORDER = ['ANTA_CRF', 'ANTA_GF', 'ANTA_LF', 'ANTA_RF', 'AURY_CRF', 'AURY_GF', 'AURY_LF', 'AURY_RF', 'CHAMERA2', 'CHAMERA3', 'DADRI_CRF', 'DADRI_GF', 'DADRI_LF', 'DADRI_RF', 'DADRT2', 'DHAULIGNGA', 'DULHASTI', 'GADARWARA', 'GANDHAR_APM', 'GANDHAR_CRF', 'GANDHAR_NAPM', 'GANDHAR_RLNG', 'JHAJJAR', 'KAPS', 'KAPS_34', 'KAWAS_APM', 'KAWAS_CRF', 'KAWAS_LF', 'KAWAS_NAPM', 'KAWAS_RLNG', 'KHARGONE', 'KHSTPP_II', 'KISHANGANGA', 'KOLDAM', 'KOTESHWR', 'KSTPS12', 'KSTPS7', 'LARA', 'MOUDA1', 'MOUDA2', 'NAPP', 'NJPC', 'PARBATI2', 'PARBATI3', 'RAMPUR', 'RAPPC', 'RAPPD', 'RIHAND1', 'RIHAND2', 'RIHAND3', 'SASAN', 'SEWA2', 'SINGRAULI', 'SINGRAULI_HYDRO', 'SIPAT1', 'SIPAT2', 'SOLAPUR', 'SSP', 'SUBANSIRI_LOWER', 'TANDA2', 'TAPS_II', 'TEHRI', 'UNCHAHAR1', 'UNCHAHAR2', 'UNCHAHAR3', 'UNCHAHAR4', 'URI2', 'VSTPS1', 'VSTPS2', 'VSTPS3', 'VSTPS4', 'VSTPS5', 'GRAND TOTAL', 'GADARWARA', 'KHARGONE', 'KSTPS12', 'KSTPS7', 'LARA', 'MOUDA1', 'MOUDA2', 'SIPAT1', 'SIPAT2', 'SOLAPUR', 'VSTPS1', 'VSTPS2', 'VSTPS3', 'VSTPS4', 'VSTPS5', 'GRAND TOTAL', 'ABCREPL_BHDL2', 'APMPL_BHDL', 'ARP1PL_BKN', 'ASIPL_BARANDA', 'AVAADA_AGAR_RUMS_S', 'AXPPL_FTG3', 'Arinsun_RUMS', 'Athena_RUMS', 'BEEMPOW_AGAR_RUMS_S', 'CPTTNPL', 'DVC', 'GADAG_GreenInfra_W', 'GADAG_VENA_S', 'GADAG_VENA_W', 'JIPL', 'JPL2', 'KHURJA_STPP', 'MEJA', 'MEL_U1_MPPMCL', 'MEL_U2_MPPMCL', 'JHABUA_IPP', 'JNSTPP', 'LANCO_AMK', 'MATATILLA-NR', 'MB POWER', 'RAJGHAT_MPSEB', 'RIHAND HYD-NR', 'TORRENT_GUVNL', 'M_WR_2020_04', 'Mahindra_RUMS', 'NTPCREL1_SJPR_RUMS_S', 'NTPCREL_PSS2_KPS2_S', 'NTPC_REL_SJPR_RUMS_S', 'NUPPL_UP_1', 'RAJGHAT_6.8MW', 'RKM_POWER_IB', 'RKM_POWER_T', 'RSEJ3PL_FTG2', 'SKS_Raigarh', 'TALTUTAI_SJPR_RUMS_S', 'TPSOURY_BRVD_NMCH_SI', 'TPSOURY_KWAI_NMCH_SI', 'TRN_ENERGY', 'Tuticorin_JSWRenew_W', 'BLA_POWER_I&II', 'JP_BINA_TPP', 'AMK_III', 'SGT', 'SGT_II', 'STP-IV', 'SSTPP-I', 'SSTPP-II', 'WIND', 'SOLAR', 'Small HYDRO', 'BIO_GAS', 'BIOMASS', 'MSW', 'GANDHI-SAGAR HPS', 'PENCH HPS', 'BARGI HPS', ' TONS HPS BSR-1', 'BIRSINGHPUR HPS', 'RAJGHAT HPS', 'DEVLOND HPS BSR-3', 'SILPARA HPS BSR-2', 'Bargi LBCPH', 'Jhinha BSR-4', 'Madikheda', 'ISP', 'OSP']

ENTT2_MODE = {'ANTA_CRF': 'OPEN_ENT', 'ANTA_GF': 'OPEN_ENT', 'ANTA_LF': 'OPEN_ENT', 'ANTA_RF': 'OPEN_ENT', 'AURY_CRF': 'OPEN_ENT', 'AURY_GF': 'OPEN_ENT', 'AURY_LF': 'OPEN_ENT', 'AURY_RF': 'OPEN_ENT', 'CHAMERA2': 'ENT', 'CHAMERA3': 'ENT', 'DADRI_CRF': 'OPEN_ENT', 'DADRI_GF': 'OPEN_ENT', 'DADRI_LF': 'OPEN_ENT', 'DADRI_RF': 'OPEN_ENT', 'DADRT2': 'ONBAR_ENT', 'DHAULIGNGA': 'ENT', 'DULHASTI': 'ENT', 'GADARWARA': 'ONBAR_ENT', 'GANDHAR_APM': 'OPEN_ENT', 'GANDHAR_CRF': 'OPEN_ENT', 'GANDHAR_NAPM': 'OPEN_ENT', 'GANDHAR_RLNG': 'OPEN_ENT', 'JHAJJAR': 'ONBAR_ENT', 'KAPS': 'ENT', 'KAPS_34': 'ENT', 'KAWAS_APM': 'OPEN_ENT', 'KAWAS_CRF': 'OPEN_ENT', 'KAWAS_LF': 'OPEN_ENT', 'KAWAS_NAPM': 'OPEN_ENT', 'KAWAS_RLNG': 'OPEN_ENT', 'KHARGONE': 'ONBAR_ENT', 'KHSTPP_II': 'ONBAR_ENT', 'KISHANGANGA': 'ENT', 'KOLDAM': 'ENT', 'KOTESHWR': 'ENT', 'KSTPS12': 'ONBAR_ENT', 'KSTPS7': 'ONBAR_ENT', 'LARA': 'ONBAR_ENT', 'MOUDA1': 'ONBAR_ENT', 'MOUDA2': 'ONBAR_ENT', 'NAPP': 'ENT', 'NJPC': 'ENT', 'PARBATI2': 'ENT', 'PARBATI3': 'ENT', 'RAMPUR': 'ENT', 'RAPPC': 'ENT', 'RAPPD': 'ENT', 'RIHAND1': 'ONBAR_ENT', 'RIHAND2': 'ONBAR_ENT', 'RIHAND3': 'ONBAR_ENT', 'SASAN': 'ONBAR_ENT', 'SEWA2': 'ENT', 'SINGRAULI': 'ONBAR_ENT', 'SINGRAULI_HYDRO': 'ENT', 'SIPAT1': 'ONBAR_ENT', 'SIPAT2': 'ONBAR_ENT', 'SOLAPUR': 'ONBAR_ENT', 'SSP': 'ENT', 'SUBANSIRI_LOWER': 'ENT', 'TANDA2': 'ONBAR_ENT', 'TAPS_II': 'ENT', 'TEHRI': 'ENT', 'UNCHAHAR1': 'ONBAR_ENT', 'UNCHAHAR2': 'ONBAR_ENT', 'UNCHAHAR3': 'ONBAR_ENT', 'UNCHAHAR4': 'ONBAR_ENT', 'URI2': 'ENT', 'VSTPS1': 'ONBAR_ENT', 'VSTPS2': 'ONBAR_ENT', 'VSTPS3': 'ONBAR_ENT', 'VSTPS4': 'ONBAR_ENT', 'VSTPS5': 'ONBAR_ENT', 'ABCREPL_BHDL2': 'DC', 'APMPL_BHDL': 'DC', 'ARP1PL_BKN': 'DC', 'ASIPL_BARANDA': 'DC', 'AVAADA_AGAR_RUMS_S': 'DC', 'AXPPL_FTG3': 'DC', 'Arinsun_RUMS': 'DC', 'Athena_RUMS': 'DC', 'BEEMPOW_AGAR_RUMS_S': 'DC', 'CPTTNPL': 'DC', 'DVC': 'OnBar', 'GADAG_GreenInfra_W': 'DC', 'GADAG_VENA_S': 'DC', 'GADAG_VENA_W': 'DC', 'JIPL': 'OnBar', 'JPL2': 'OnBar', 'KHURJA_STPP': 'OnBar', 'MEJA': 'OnBar', 'MEL_U1_MPPMCL': 'OnBar', 'MEL_U2_MPPMCL': 'OnBar', 'JHABUA_IPP': 'OnBar', 'JNSTPP': 'OnBar', 'LANCO_AMK': 'Deemed Capability', 'MATATILLA-NR': 'DC', 'MB POWER': 'OnBar', 'RAJGHAT_MPSEB': 'DC', 'RIHAND HYD-NR': 'DC', 'TORRENT_GUVNL': 'Deemed Capability', 'M_WR_2020_04': 'DC', 'Mahindra_RUMS': 'DC', 'NTPCREL1_SJPR_RUMS_S': 'DC', 'NTPCREL_PSS2_KPS2_S': 'DC', 'NTPC_REL_SJPR_RUMS_S': 'DC', 'NUPPL_UP_1': 'OnBar', 'RAJGHAT_6.8MW': 'DC', 'RKM_POWER_IB': 'OnBar', 'RKM_POWER_T': 'OnBar', 'RSEJ3PL_FTG2': 'DC', 'SKS_Raigarh': 'OnBar', 'TALTUTAI_SJPR_RUMS_S': 'DC', 'TPSOURY_BRVD_NMCH_SI': 'DC', 'TPSOURY_KWAI_NMCH_SI': 'DC', 'TRN_ENERGY': 'OnBar', 'Tuticorin_JSWRenew_W': 'DC', 'BLA_POWER_I&II': 'OnBar', 'JP_BINA_TPP': 'OnBar'}

SSGS_SEZ_NAMES = ["AMK_III", "SGT", "SGT_II", "STP-IV", "SSTPP-I", "SSTPP-II"]
BID_PRICE_POINTS = [
    2899, 2900, 2928, 2929, 2970, 2971, 3110, 3111, 3226, 3227, 3280, 3281, 3295,
    3296, 3313, 3314, 3563, 3564, 3640, 3641, 3757, 3758, 4040, 4041, 4298, 4299, 10000,
]
BID_MARGIN = 0.15
KEEP_SHEETS = {
    "Name corrector",
    "Deamnd-MP",
    "ENTT-2",
    "RE",
    "Hydel",
    "Hydel ReSchedule",
    "Entt",
    "Scheduling",
    "BD",
    "Bid Sheet",
    "IEX",
}


def prune_unlinked_sheets(wb: Workbook) -> None:
    # Remove sheets that are not part of the main logic/output.
    for name in list(wb.sheetnames):
        if name not in KEEP_SHEETS:
            del wb[name]

    # If IEX is not referenced by any remaining sheet, remove it.
    if "IEX" in wb.sheetnames:
        ref_re = re.compile(r"'([^']+)'!|([A-Za-z0-9_ ]+)!", re.IGNORECASE)
        referenced = set()
        for sname in wb.sheetnames:
            ws = wb[sname]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        for m in ref_re.finditer(v):
                            ref = m.group(1) or m.group(2)
                            if ref in wb.sheetnames and ref != sname:
                                referenced.add(ref)
        if "IEX" not in referenced:
            del wb["IEX"]


def reorder_sheets(wb: Workbook, preferred_order: list[str]) -> None:
    # Reorder sheets: preferred first, then remaining in existing order.
    name_to_sheet = {ws.title: ws for ws in wb.worksheets}
    ordered = []
    for name in preferred_order:
        ws = name_to_sheet.get(name)
        if ws is not None:
            ordered.append(ws)
    for ws in wb.worksheets:
        if ws not in ordered:
            ordered.append(ws)
    wb._sheets = ordered


def format_bid_date_label(date_val: object) -> str:
    if hasattr(date_val, "strftime"):
        return date_val.strftime("%d.%m.%y")
    return str(date_val or "").strip()


def setup_bid_sheet_layout(
    ws_bids: openpyxl.worksheet.worksheet.Worksheet,
    date_val: object,
    bid_prices: list[int],
    bid_margin: float,
    demand_ez_present: bool,
) -> None:
    start_col = 3
    header_merge_end_col = 15  # Column O
    last_col = start_col + len(bid_prices) - 1

    # Header merges (match sample layout)
    ws_bids.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws_bids.merge_cells(start_row=1, start_column=3, end_row=1, end_column=header_merge_end_col)
    ws_bids.merge_cells(start_row=2, start_column=3, end_row=2, end_column=header_merge_end_col)
    ws_bids.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)
    ws_bids.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)

    date_label = format_bid_date_label(date_val)
    if date_label:
        ws_bids.cell(1, 1, f"IEX Transaction For {date_label}")
    else:
        ws_bids.cell(1, 1, "IEX Transaction For")

    if demand_ez_present:
        ws_bids.cell(1, 3, '=CONCATENATE("BID FOR ",TEXT(\'Demand-EZ\'!E2,("dd-mmm-yy")))')
    elif date_label:
        ws_bids.cell(1, 3, f"BID FOR {date_label}")
    else:
        ws_bids.cell(1, 3, "BID FOR")

    ws_bids.cell(2, 1, "S.No.")
    ws_bids.cell(2, 2, "Period")
    ws_bids.cell(2, 3, "BID RATE  /  BID VOLUME")

    # Bid margin (O3)
    margin_col = 15
    ws_bids.cell(3, margin_col, bid_margin).number_format = "0.00"

    # Price points row (row 4) and /1000 row (row 5)
    for idx, price in enumerate(bid_prices):
        col = start_col + idx
        col_letter = get_column_letter(col)
        ws_bids.cell(4, col, price).number_format = "0.00"
        ws_bids.cell(5, col, f"={col_letter}4/1000").number_format = "0.00"

    # Row heights
    row_heights = {1: 17.25, 2: 54.0, 3: 22.5, 4: 19.5, 5: 19.5, 6: 21.75}
    for r, h in row_heights.items():
        ws_bids.row_dimensions[r].height = h

    # Column widths (sample layout)
    bid_widths = {
        "A": 10.57, "B": 16.0, "C": 10.0, "D": 13.0, "E": 13.0, "F": 13.0, "G": 13.0,
        "H": 13.0, "I": 13.0, "J": 13.0, "K": 13.0, "L": 13.0, "M": 13.0, "N": 13.0,
        "O": 10.43, "P": 12.86, "R": 9.14, "S": 13.0, "T": 13.0, "U": 13.0, "V": 13.0,
        "W": 13.0, "X": 13.0, "Y": 13.0, "Z": 13.0, "AA": 13.0, "AB": 13.0, "AC": 13.0,
    }
    for col_letter, width in bid_widths.items():
        ws_bids.column_dimensions[col_letter].width = width

    # Basic styles (borders + alignment)
    header_font = Font(bold=True, color="000000")
    data_font = Font(bold=True, color="000000")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in range(1, 6):
        for c in range(1, last_col + 1):
            cell = ws_bids.cell(r, c)
            if cell.value in (None, ""):
                continue
            cell.font = header_font
            cell.alignment = center
            cell.border = border

    # Apply grid styling for data area
    for r in range(6, 102):
        for c in range(1, last_col + 1):
            cell = ws_bids.cell(r, c)
            cell.font = data_font
            cell.alignment = center
            cell.border = border


def clean_station(name: object) -> str:
    if name is None:
        return ""
    text = str(name).strip()
    text = re.sub(r"\s+", " ", text)
    return ACTIVE_NAME_MAP.get(text, text)


def safe_float(value: object) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def parse_number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return 0.0
    s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return 0.0


def parse_iex_pdf(path: Path) -> dict[str, dict[str, float]]:
    import pdfplumber

    line_re = re.compile(r"(\d{2}:\d{2}\s*-\s*\d{2}:\d{2})\s+([\d,.\-]+)\s+([\d,.]+)\s+([\d,.\-]+)")
    rows: list[tuple[str, float, float, float]] = []

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                if "Period Qty" in line or "Daily Trade Report" in line:
                    continue
                matches = line_re.findall(line)
                if not matches:
                    continue
                for period, qty, rate, amt in matches:
                    rows.append((
                        re.sub(r"\\s+", " ", period.strip()),
                        parse_number(qty),
                        parse_number(rate),
                        parse_number(amt),
                    ))

    if not rows:
        raise ValueError(f"Could not parse IEX PDF: {path.name}")

    out: dict[str, dict[str, float]] = {}
    for period, qty, rate, amt in rows:
        out[period] = {"qty": qty, "rate": rate, "amount": amt}
    return out


def norm_mode(value: object) -> str:
    text = re.sub(r"[\s_]+", "", str(value or "").strip().upper())
    if text == "COMBINEDENT":
        return "COMBINEDENT"
    if text == "ONBARENT":
        return "ONBARENT"
    if text == "OPENENT":
        return "OPENENT"
    if text == "OFFBARENT":
        return "OFFBARENT"
    if text in {"TOTALENT", "GRANDTOTALENT"}:
        return "TOTALENT"
    if text == "ENT":
        return "ENT"
    if text == "DC":
        return "DC"
    if text == "ONBAR":
        return "ONBAR"
    if text == "DECLAREDCAPABILITY":
        return "DECLARED"
    if text == "DEEMEDCAPABILITY":
        return "DEEMED"
    return text


def parse_demand_file(path: Path) -> dict[int, dict[str, object]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    def norm(text: object) -> str:
        return re.sub(r"\s+", " ", str(text or "").strip().lower())

    def block_to_time_window(block: int) -> tuple[str, str]:
        start = (block - 1) * 15
        end = block * 15
        return f"{(start // 60) % 24:02d}:{start % 60:02d}:00", f"{(end // 60) % 24:02d}:{end % 60:02d}:00"

    header_row = None
    for r in range(1, min(40, ws.max_row + 1)):
        row_vals = [norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        has_time_or_block = any("time block" in v for v in row_vals) or any("block" in v and "time" in v for v in row_vals)
        has_restricted = any("restricted" in v and "un-restricted" not in v and "unrestricted" not in v for v in row_vals)
        if has_time_or_block and has_restricted:
            header_row = r
            break

    if header_row is not None:
        restricted_col = None
        block_col = None
        from_col = None
        to_col = None

        for c in range(1, ws.max_column + 1):
            cell = norm(ws.cell(header_row, c).value)
            if "restricted" in cell and "un-restricted" not in cell and "unrestricted" not in cell:
                restricted_col = c
            if ("sl.no" in cell or "block" in cell) and "time block" not in cell:
                block_col = c

        for rr in [header_row, header_row + 1, header_row + 2]:
            if rr > ws.max_row:
                continue
            for c in range(1, ws.max_column + 1):
                cell = norm(ws.cell(rr, c).value)
                if cell == "from" and from_col is None:
                    from_col = c
                if cell == "to" and to_col is None:
                    to_col = c

        if all([restricted_col, block_col, from_col, to_col]):
            out: dict[int, dict[str, object]] = {}
            for r in range(header_row + 2, ws.max_row + 1):
                block = ws.cell(r, block_col).value
                if isinstance(block, (int, float)):
                    block = int(block)
                    if 1 <= block <= 96:
                        out[block] = {
                            "from": ws.cell(r, from_col).value,
                            "to": ws.cell(r, to_col).value,
                            "demand": safe_float(ws.cell(r, restricted_col).value),
                        }
            if out:
                return out

    # Fallback format: Date, From Block, To Block, ..., value, Unit
    alt_header = None
    for r in range(1, min(20, ws.max_row + 1)):
        row_vals = [norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if any(v == "from block" for v in row_vals) and any(v == "value" for v in row_vals):
            alt_header = r
            break

    if alt_header is not None:
        from_block_col = None
        value_col = None
        for c in range(1, ws.max_column + 1):
            cell = norm(ws.cell(alt_header, c).value)
            if cell == "from block":
                from_block_col = c
            if cell == "value":
                value_col = c
        if from_block_col and value_col:
            out: dict[int, dict[str, object]] = {}
            for r in range(alt_header + 1, ws.max_row + 1):
                block = ws.cell(r, from_block_col).value
                if not isinstance(block, (int, float)):
                    continue
                b = int(block)
                if not (1 <= b <= 96):
                    continue
                from_t, to_t = block_to_time_window(b)
                out[b] = {
                    "from": from_t,
                    "to": to_t,
                    "demand": safe_float(ws.cell(r, value_col).value),
                }
            if out:
                return out

    raise ValueError(f"Could not detect demand header in: {path.name}")


def add_station_block(target: dict[int, dict[str, float]], block: int, station: str, value: float) -> None:
    if not station:
        return
    station = clean_station(station)
    if ALLOWED_STATIONS and station not in ALLOWED_STATIONS:
        return
    target[block][station] += value


def parse_entitlement_file(path: Path) -> dict[int, dict[str, float]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ent = defaultdict(lambda: defaultdict(float))

    def norm(text: object) -> str:
        return re.sub(r"\s+", " ", str(text or "").strip().upper())

    def find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
        for r in range(1, min(12, ws.max_row + 1)):
            if norm(ws.cell(r, 2).value) == "TIME BLOCK":
                return r
        return None

    def find_data_row(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int) -> int | None:
        for r in range(header_row + 1, ws.max_row + 1):
            block = ws.cell(r, 2).value
            if isinstance(block, (int, float)) and 1 <= int(block) <= 96:
                return r
        return None

    def is_grand_total(station: object, mode: object) -> bool:
        t = f"{norm(station)} {norm(mode)}"
        return "GRAND" in t and "TOTAL" in t

    def scan_modes(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int, data_row: int, col: int) -> str:
        vals = [norm(ws.cell(r, col).value) for r in range(header_row, data_row)]
        return " | ".join(v for v in vals if v)

    def add_sheet_by_mode(sheet_name: str, mode_keyword: str) -> None:
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        header_row = find_header_row(ws)
        if header_row is None:
            return
        data_row = find_data_row(ws, header_row)
        if data_row is None:
            return

        station_names: list[str] = []
        last_station = ""
        for c in range(1, ws.max_column + 1):
            raw = str(ws.cell(header_row, c).value or "").strip()
            if raw:
                last_station = raw
            station_names.append(last_station)

        selected_cols: list[int] = []
        for c in range(5, ws.max_column + 1):
            header_modes = scan_modes(ws, header_row, data_row, c)
            if mode_keyword in header_modes and not is_grand_total(station_names[c - 1], header_modes):
                selected_cols.append(c)

        for r in range(data_row, ws.max_row + 1):
            block = ws.cell(r, 2).value
            if not isinstance(block, (int, float)):
                continue
            b = int(block)
            if not (1 <= b <= 96):
                continue
            for c in selected_cols:
                add_station_block(ent, b, station_names[c - 1], safe_float(ws.cell(r, c).value))

    # ISGS + ISGS NTPC RAJ SOLAR: use TOTAL_ENT only
    add_sheet_by_mode("ISGS", "TOTAL_ENT")
    add_sheet_by_mode("ISGS NTPC RAJ SOLAR", "TOTAL_ENT")

    # REMC & IPP: use DC columns
    add_sheet_by_mode("REMC & IPP", "DC")

    # SSGS: use entitlement columns (AMK_III, SGT, SGT_II, STP-IV, SSTPP-I, SSTPP-II)
    if "SSGS" in wb.sheetnames:
        ws = wb["SSGS"]
        header_row = find_header_row(ws)
        if header_row is not None:
            data_row = find_data_row(ws, header_row)
            if data_row is not None:
                targets = ["AMK_III", "SGT", "SGT_II", "STP-IV", "SSTPP-I", "SSTPP-II"]
                col_by_station: dict[str, int] = {}
                for idx, target in enumerate(targets, start=5):
                    col_by_station[target] = idx

                for r in range(data_row, ws.max_row + 1):
                    block = ws.cell(r, 2).value
                    if not isinstance(block, (int, float)):
                        continue
                    b = int(block)
                    if not (1 <= b <= 96):
                        continue
                    for station, c in col_by_station.items():
                        add_station_block(ent, b, station, safe_float(ws.cell(r, c).value))

    return ent


def copy_sheet_values(src_ws: openpyxl.worksheet.worksheet.Worksheet, dst_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for r in range(1, src_ws.max_row + 1):
        for c in range(1, src_ws.max_column + 1):
            dst_ws.cell(r, c, src_ws.cell(r, c).value)


def copy_sheet_full(src_ws: openpyxl.worksheet.worksheet.Worksheet, dst_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for r in range(1, src_ws.max_row + 1):
        for c in range(1, src_ws.max_column + 1):
            s = src_ws.cell(r, c)
            d = dst_ws.cell(r, c, s.value)
            if s.has_style:
                d.font = copy(s.font)
                d.fill = copy(s.fill)
                d.border = copy(s.border)
                d.alignment = copy(s.alignment)
                d.number_format = s.number_format
                d.protection = copy(s.protection)
    for key, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[key].width = dim.width
    for key, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[key].height = dim.height
    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))


def find_timeblock_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
    for r in range(1, min(20, ws.max_row + 1)):
        if norm_key(ws.cell(r, 2).value) == "time block":
            return r
    return None


def find_timeblock_data_row(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int) -> int | None:
    for r in range(header_row + 1, ws.max_row + 1):
        block = ws.cell(r, 2).value
        if isinstance(block, (int, float)) and 1 <= int(block) <= 96:
            return r
    return None


def add_total_column(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int,
                     data_row: int, start_col: int = 5) -> dict[int, float]:
    total_col = ws.max_column + 1
    ws.cell(header_row, total_col, "Total")
    totals: dict[int, float] = {}
    for r in range(data_row, ws.max_row + 1):
        block = ws.cell(r, 2).value
        if not isinstance(block, (int, float)):
            continue
        b = int(block)
        if not (1 <= b <= 96):
            continue
        s = 0.0
        for c in range(start_col, total_col):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                s += float(v)
        ws.cell(r, total_col, s)
        totals[b] = s
    return totals


def autofit_header_widths(wb: openpyxl.Workbook, header_rows: tuple[int, ...] = (1, 2, 3, 4),
                          min_width: float = 8.0, max_width: float = 60.0) -> None:
    for ws in wb.worksheets:
        max_col = ws.max_column or 0
        for c in range(1, max_col + 1):
            col_letter = get_column_letter(c)
            dim = ws.column_dimensions[col_letter]
            if dim.width is not None:
                continue
            max_len = 0
            for r in header_rows:
                val = ws.cell(r, c).value
                if val in (None, ""):
                    continue
                s = str(val)
                if "\n" in s:
                    part_len = max(len(p) for p in s.splitlines())
                else:
                    part_len = len(s)
                if part_len > max_len:
                    max_len = part_len
            if max_len:
                dim.width = min(max_width, max(min_width, max_len + 2))


def parse_mod_merit(path: Path) -> list[tuple[str, float]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = None
    for name in wb.sheetnames:
        n = re.sub(r"\s+", " ", str(name).strip().lower())
        if n in {"periphery 1", "periphery-1", "periphery1"}:
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    def norm(text: object) -> str:
        return re.sub(r"\s+", " ", str(text or "").strip().lower())

    header_row = None
    for r in range(1, min(40, ws.max_row + 1)):
        vals = [norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if any("name of generating station" in v for v in vals):
            header_row = r
            break
    if header_row is None:
        raise ValueError("Could not detect MOD header row")

    name_col = None
    rate_col = None
    alt_rate_col = None
    eff_col = None
    for c in range(1, ws.max_column + 1):
        v = norm(ws.cell(header_row, c).value)
        if "name of generating station" in v:
            name_col = c
        if ("variable" in v and "rate" in v) or "rs./kwh" in v or "rs/kwh" in v:
            rate_col = c
    if not name_col or not rate_col:
        raise ValueError("Could not find MOD name/rate columns")

    merit: list[tuple[str, float]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        rate = ws.cell(r, rate_col).value
        if name in (None, ""):
            continue
        try:
            rate_val = float(rate)
        except Exception:
            continue
        merit.append((clean_station(name), rate_val))

    # First occurrence per station
    unique: dict[str, float] = {}
    for station, rate in merit:
        if station and station not in unique:
            unique[station] = rate

    return sorted(unique.items(), key=lambda x: x[1])


def parse_mod_rates_excl_urs3(path: Path) -> dict[str, float]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if str(name).strip().upper() == "EXCL URS 3":
            ws = wb[name]
            break
    if ws is None:
        return {}

    def norm(text: object) -> str:
        return re.sub(r"\s+", " ", str(text or "").strip().lower())

    header_row = None
    for r in range(1, min(50, ws.max_row + 1)):
        vals = [norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if any("name of generating station" in v for v in vals) and any("effective" in v and "rs" in v for v in vals):
            header_row = r
            break
    if header_row is None:
        return {}

    name_col = None
    rate_col = None
    for c in range(1, ws.max_column + 1):
        v = norm(ws.cell(header_row, c).value)
        if "name of generating station" in v:
            name_col = c
        if "effective" in v and "rs" in v:
            rate_col = c
        if "variable cost" in v and "losses" in v:
            alt_rate_col = c
        if "effective energy charges" in v:
            eff_col = c
    if not name_col or not rate_col:
        return {}

    rates: dict[str, float] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        rate = ws.cell(r, rate_col).value
        if name in (None, ""):
            continue
        rate_val = None
        if rate not in (None, ""):
            try:
                rate_val = float(rate)
            except Exception:
                rate_val = None
        if rate_val is None and alt_rate_col:
            alt = ws.cell(r, alt_rate_col).value
            try:
                rate_val = float(alt)
            except Exception:
                rate_val = None
        if rate_val is None and eff_col:
            alt = ws.cell(r, eff_col).value
            try:
                rate_val = float(alt)
            except Exception:
                rate_val = None
        if rate_val is None:
            continue
        rates[norm(name)] = rate_val
    return rates


def demand_series_equal(a: dict[int, dict[str, object]], b: dict[int, dict[str, object]]) -> bool:
    for block in range(1, 97):
        av = safe_float(a.get(block, {}).get("demand"))
        bv = safe_float(b.get(block, {}).get("demand"))
        if abs(av - bv) > 1e-9:
            return False
    return True


def build_outputs(
    demand_cz_path: Path | None,
    demand_ez_path: Path | None,
    demand_wz_path: Path | None,
    iex_pdf_path: Path | None,
    entitlement_path: Path,
    mod_rate_path: Path,
    output_path: Path,
) -> None:
    demand_cz = parse_demand_file(demand_cz_path) if demand_cz_path else {}
    demand_ez = parse_demand_file(demand_ez_path) if demand_ez_path else {}
    demand_wz = parse_demand_file(demand_wz_path) if demand_wz_path else {}
    iex_data = parse_iex_pdf(iex_pdf_path) if iex_pdf_path else {}

    ent: dict[int, dict[str, float]] = {}
    merit_all = parse_mod_merit(mod_rate_path)
    vc_by_station = {station: vc for station, vc in merit_all}

    # MOD-based merit list (used for rank columns)
    ent_src = openpyxl.load_workbook(entitlement_path, data_only=True)
    mod_stations = list(RANKED_STATIONS)

    # REMC & IPP stations (append after merit list)
    dc_stations: list[str] = []
    if "REMC & IPP" in ent_src.sheetnames:
        ws_dc = ent_src["REMC & IPP"]
        for c in range(5, ws_dc.max_column + 1):
            name = ws_dc.cell(1, c).value
            if name:
                dc = clean_station(name)
                if dc and dc not in dc_stations:
                    dc_stations.append(dc)

    # All stations in entitlement, keep DC at the end
    stations = mod_stations + [None, None] + [s for s in dc_stations if s not in mod_stations]
    vcs = [vc_by_station.get(s, 0.0) for s in stations]
    n = len(stations)

    wb = Workbook()
    wb.remove(wb.active)


    # Name corrector (always build from internal map)
    ws_nc = wb.create_sheet("Name corrector")
    ws_nc.cell(2, 3, "Station names")
    ws_nc.cell(2, 4, "New names")
    ws_nc.cell(2, 5, "Ent")
    ws_nc.cell(2, 6, "MOD Rates")
    ws_nc.cell(2, 7, "Rank")
    seen_orig = set()
    row = 3
    for orig, code in NAME_MAP.items():
        key = norm_key(orig)
        if key in seen_orig:
            continue
        seen_orig.add(key)
        ws_nc.cell(row, 3, orig)
        ws_nc.cell(row, 4, code)
        row += 1

    # Style Name corrector header
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True, color="000000")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(3, 8):
        cell = ws_nc.cell(2, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Align data rows for readability
    data_align = Alignment(horizontal="left", vertical="center")
    for r in range(3, ws_nc.max_row + 1):
        for c in range(3, 8):
            ws_nc.cell(r, c).alignment = data_align

    # Build Ent category from entitlement sheets
    hydel_codes = set(HYDEL_CODES)
    ssgs_codes = set(SSGS_CODES)
    if "Intra State Hydel Schedule" in ent_src.sheetnames:
        ws_hyd_src = ent_src["Intra State Hydel Schedule"]
        for c in range(5, ws_hyd_src.max_column + 1):
            name = ws_hyd_src.cell(1, c).value
            if name:
                hydel_codes.add(norm_key(name))
    if "SSGS" in ent_src.sheetnames:
        ws_ssgs_src = ent_src["SSGS"]
        for c in range(5, ws_ssgs_src.max_column + 1):
            name = ws_ssgs_src.cell(2, c).value
            if name:
                ssgs_codes.add(norm_key(name))

    # MOD rates and ranks from EXCL URS 3
    mod_rate_map = parse_mod_rates_excl_urs3(mod_rate_path)
    # rank by MOD rate descending (higher rate => higher rank)
    sorted_rates = sorted(mod_rate_map.items(), key=lambda x: x[1])
    rank_map = {name: idx + 1 for idx, (name, _) in enumerate(sorted_rates)}

    def norm_name(text: object) -> str:
        return re.sub(r"\s+", " ", str(text or "").strip().lower())

    # Fill rows based on internal station list
    for r in range(3, ws_nc.max_row + 1):
        orig = ws_nc.cell(r, 3).value
        code = ws_nc.cell(r, 4).value
        if orig in (None, "") and code in (None, ""):
            continue
        # Preserve Ent column if provided; only fill when blank
        if ws_nc.cell(r, 5).value in (None, ""):
            code_key = norm_key(code)
            if code_key in hydel_codes:
                ws_nc.cell(r, 5, "Hydel")
            elif code_key in ssgs_codes:
                ws_nc.cell(r, 5, "SSGS")
            else:
                ws_nc.cell(r, 5, "ISGS")

        rate = mod_rate_map.get(norm_name(orig))
        if rate is not None:
            ws_nc.cell(r, 6, rate)
        else:
            ws_nc.cell(r, 6, None)
        ws_nc.cell(r, 7, None)

    # Rank using fixed station order so Entt/Scheduling/BD align with RANKED_STATIONS
    ranked_index = {norm_key(name): idx + 1 for idx, name in enumerate(RANKED_STATIONS)}
    for r in range(3, ws_nc.max_row + 1):
        code = ws_nc.cell(r, 4).value
        if code in (None, ""):
            continue
        rank = ranked_index.get(norm_key(code))
        ws_nc.cell(r, 7, rank if rank is not None else None)

    # Demand sheets (raw copy of input)
    if demand_ez_path:
        src = openpyxl.load_workbook(demand_ez_path, data_only=True)
        ws = wb.create_sheet("Demand-EZ")
        copy_sheet_full(src[src.sheetnames[0]], ws)
    if demand_wz_path:
        src = openpyxl.load_workbook(demand_wz_path, data_only=True)
        ws = wb.create_sheet("Demand-WZ")
        copy_sheet_full(src[src.sheetnames[0]], ws)
    if demand_cz_path:
        src = openpyxl.load_workbook(demand_cz_path, data_only=True)
        ws = wb.create_sheet("Demand-CZ")
        copy_sheet_full(src[src.sheetnames[0]], ws)

    # Copy ENTT-1 from entitlement if available (used by Hydel ReSchedule formulas)
    ws_entt1 = None
    if "ENTT-1" in ent_src.sheetnames:
        ws_entt1 = wb.create_sheet("ENTT-1")
        copy_sheet_full(ent_src["ENTT-1"], ws_entt1)

    # Entitlement source workbook already loaded above (used directly for ENTT-2)

    # ENTT-2 (selected single-column view)
    ws_entt2 = wb.create_sheet("ENTT-2")
    ws_entt2.cell(1, 1, "Date")
    ws_entt2.cell(1, 2, "Time Block")
    ws_entt2.cell(1, 3, "from")
    ws_entt2.cell(1, 4, "to")
    ws_entt2.cell(2, 1, None)
    ws_entt2.cell(2, 2, None)
    ws_entt2.cell(2, 3, None)
    ws_entt2.cell(2, 4, None)
    ws_entt2.cell(3, 1, None)
    ws_entt2.cell(3, 2, None)
    ws_entt2.cell(3, 3, None)
    ws_entt2.cell(3, 4, None)

    # Build indices from entitlement source sheets
    ws_isgs = ent_src["ISGS"] if "ISGS" in ent_src.sheetnames else None
    ws_ntpc = ent_src["ISGS NTPC RAJ SOLAR"] if "ISGS NTPC RAJ SOLAR" in ent_src.sheetnames else None
    ws_remc = ent_src["REMC & IPP"] if "REMC & IPP" in ent_src.sheetnames else None
    ws_ssgs = ent_src["SSGS"] if "SSGS" in ent_src.sheetnames else None
    ws_re_src = ent_src["Intra State RE"] if "Intra State RE" in ent_src.sheetnames else None
    ws_hyd_src = ent_src["Intra State Hydel Schedule"] if "Intra State Hydel Schedule" in ent_src.sheetnames else None

    def build_index(ws, name_row, mode_row, start_col):
        cols_by_station: dict[str, list[int]] = defaultdict(list)
        mode_by_col: dict[int, object] = {}
        if ws is None:
            return cols_by_station, mode_by_col
        for c in range(start_col, ws.max_column + 1):
            name = ws.cell(name_row, c).value
            if name not in (None, ""):
                cols_by_station[str(name)].append(c)
                if mode_row:
                    mode_by_col[c] = ws.cell(mode_row, c).value
        return cols_by_station, mode_by_col

    isgs_cols, isgs_mode = build_index(ws_isgs, 1, 3, 5)
    ntpc_cols, ntpc_mode = build_index(ws_ntpc, 1, 3, 5)
    remc_cols, remc_mode = build_index(ws_remc, 1, 2, 5)
    ssgs_cols, _ = build_index(ws_ssgs, 2, None, 5)
    ssgs_sez_cols, _ = build_index(ws_ssgs, 2, None, 13)
    re_cols, _ = build_index(ws_re_src, 1, None, 5)
    hyd_cols, _ = build_index(ws_hyd_src, 1, None, 5)

    # Copy Date/Time/From/To from ISGS if available
    if ws_isgs is not None:
        for block in range(1, 97):
            r = block + 3
            src_r = block + 3
            for c in range(1, 5):
                ws_entt2.cell(r, c, ws_isgs.cell(src_r, c).value)

    # Force ENTT-2 ES:EX from SSGS M:R (columns 13..18), if available
    if ws_ssgs is not None:
        ssgs_start_col = column_index_from_string("M")
        entt2_start_col = column_index_from_string("ES")
        # Set headers for ES:EX from SSGS header row (row 2 in SSGS)
        ssgs_header_row = 2
        for offset in range(6):
            ws_entt2.cell(1, entt2_start_col + offset, ws_ssgs.cell(ssgs_header_row, ssgs_start_col + offset).value)
        for block in range(1, 97):
            r = block + 3
            for offset in range(6):
                ws_entt2.cell(r, entt2_start_col + offset, ws_ssgs.cell(r, ssgs_start_col + offset).value)

    def pick_col(cols_by_station, mode_by_col, station, desired_mode, used_key):
        cols = cols_by_station.get(station, [])
        if not cols:
            return None
        if desired_mode:
            m = norm_mode(desired_mode)
            if m == "OPENENT":
                filtered = [c for c in cols if norm_mode(mode_by_col.get(c)) == "OPENENT"]
                if not filtered:
                    filtered = [c for c in cols if norm_mode(mode_by_col.get(c)) == "COMBINEDENT"]
            elif m == "ONBARENT":
                filtered = [c for c in cols if norm_mode(mode_by_col.get(c)) == "ONBARENT"]
                if not filtered:
                    filtered = [c for c in cols if norm_mode(mode_by_col.get(c)) == "OFFBARENT"]
            else:
                filtered = [c for c in cols if norm_mode(mode_by_col.get(c)) == m]
            if filtered:
                idx = min(used_key, len(filtered) - 1)
                return filtered[idx]
        idx = min(used_key, len(cols) - 1)
        return cols[idx]

    used_count: dict[tuple[str, str, str], int] = defaultdict(int)
    sez_entt2_cols: list[int] = []

    if True:
        col = 5
        blank_cols = {78, 95, 142, 155, 175, 176, 177}
        sez_start_col = 149
        ntpc_section = False
        grand_total_seen = 0

        for station in ENTT2_ORDER:
            while col in blank_cols:
                ws_entt2.cell(1, col, "")
                col += 1

            if col == sez_start_col:
                for sez_name in SSGS_SEZ_NAMES:
                    ws_entt2.cell(1, col, None)
                    ws_entt2.cell(2, col, sez_name)
                    ws_entt2.cell(3, col, sez_name)
                    pick = pick_col(ssgs_sez_cols, {}, str(sez_name), None, 0)
                    for block in range(1, 97):
                        r = block + 3
                        src_r = block + 2
                        if ws_ssgs is not None and pick:
                            ws_entt2.cell(r, col, ws_ssgs.cell(src_r, pick).value)
                    sez_entt2_cols.append(col)
                    col += 1
            if col == 155:
                ws_entt2.cell(1, col, "")
                col += 1

            ws_entt2.cell(1, col, station)
            mode = ENTT2_MODE.get(station)
            if station == "GRAND TOTAL":
                ws_entt2.cell(2, col, "GRAND TOTAL")
                grand_total_seen += 1
                ntpc_section = grand_total_seen == 1
            elif ntpc_section and station in ntpc_cols:
                ws_entt2.cell(2, col, "RAJ_SOLAR_MPSEB")
            elif mode and norm_mode(mode) in {"OPENENT", "ONBARENT", "ENT", "TOTALENT"}:
                ws_entt2.cell(2, col, "MPSEB_Beneficiary")
            elif mode:
                ws_entt2.cell(2, col, station)
            if mode:
                ws_entt2.cell(3, col, mode)

            station_key = str(station)

            if station_key in re_cols:
                pick = re_cols[station_key][0]
                for block in range(1, 97):
                    r = block + 3
                    src_r = block + 1
                    if ws_re_src is not None:
                        ws_entt2.cell(r, col, ws_re_src.cell(src_r, pick).value)
                col += 1
                continue

            if station_key in hyd_cols:
                pick = hyd_cols[station_key][0]
                for block in range(1, 97):
                    r = block + 3
                    src_r = block + 1
                    if ws_hyd_src is not None:
                        ws_entt2.cell(r, col, ws_hyd_src.cell(src_r, pick).value)
                col += 1
                continue

            if station_key in ssgs_cols:
                pick = ssgs_cols[station_key][0]
                for block in range(1, 97):
                    r = block + 3
                    src_r = block + 2
                    if ws_ssgs is not None:
                        ws_entt2.cell(r, col, ws_ssgs.cell(src_r, pick).value)
                col += 1
                continue

            if ntpc_section and station_key in ntpc_cols:
                idx = used_count[(station_key, str(mode or ""), "ntpc")]
                pick = pick_col(ntpc_cols, ntpc_mode, station_key, mode, idx)
                used_count[(station_key, str(mode or ""), "ntpc")] += 1
                for block in range(1, 97):
                    r = block + 3
                    src_r = block + 3
                    if ws_ntpc is not None and pick:
                        ws_entt2.cell(r, col, ws_ntpc.cell(src_r, pick).value)
                col += 1
                continue

            if station_key in isgs_cols:
                idx = used_count[(station_key, str(mode or ""), "isgs")]
                pick = pick_col(isgs_cols, isgs_mode, station_key, mode, idx)
                used_count[(station_key, str(mode or ""), "isgs")] += 1
                for block in range(1, 97):
                    r = block + 3
                    src_r = block + 3
                    if ws_isgs is not None and pick:
                        ws_entt2.cell(r, col, ws_isgs.cell(src_r, pick).value)
                col += 1
                continue

            if station_key in remc_cols:
                idx = used_count[(station_key, str(mode or ""), "remc")]
                pick = pick_col(remc_cols, remc_mode, station_key, mode, idx)
                used_count[(station_key, str(mode or ""), "remc")] += 1
                for block in range(1, 97):
                    r = block + 3
                    src_r = block + 2
                    if ws_remc is not None and pick:
                        ws_entt2.cell(r, col, ws_remc.cell(src_r, pick).value)

            col += 1

        while col <= 177:
            ws_entt2.cell(1, col, "")
            col += 1

    # Style ENTT-2 header rows
    entt2_header_fill = PatternFill("solid", fgColor="D9D9D9")
    entt2_header_font = Font(bold=True, color="000000")
    entt2_header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    entt2_thin = Side(style="thin", color="000000")
    entt2_border = Border(left=entt2_thin, right=entt2_thin, top=entt2_thin, bottom=entt2_thin)
    max_col = ws_entt2.max_column
    for r in (1, 2, 3):
        for c in range(1, max_col + 1):
            cell = ws_entt2.cell(r, c)
            if cell.value in (None, "") and c > 4:
                continue
            cell.fill = entt2_header_fill
            cell.font = entt2_header_font
            cell.alignment = entt2_header_align
            cell.border = entt2_border
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        dim = ws_entt2.column_dimensions[col_letter]
        if dim.width is None or dim.width < 18:
            dim.width = 18

    # Build ENTT-2 column map for downstream sheets
    entt2_col_by_name: dict[str, int] = {}
    for c in range(1, ws_entt2.max_column + 1):
        name = ws_entt2.cell(1, c).value
        if name:
            entt2_col_by_name[str(name)] = c
    entt2_col_by_norm: dict[str, int] = {norm_key(k): v for k, v in entt2_col_by_name.items()}

    # Build entitlement map from ENTT-2 (first occurrence per station)
    entt2_first_col: dict[str, int] = {}
    for c in range(1, ws_entt2.max_column + 1):
        name = ws_entt2.cell(1, c).value
        if name and name not in entt2_first_col:
            entt2_first_col[str(name)] = c
    ent = defaultdict(lambda: defaultdict(float))
    for block in range(1, 97):
        r = block + 3
        for st in stations:
            if not st:
                continue
            col_idx = entt2_first_col.get(st)
            if col_idx:
                ent[block][st] = safe_float(ws_entt2.cell(r, col_idx).value)

    # RE sheet (compact layout referencing ENTT-2)
    ws_re = wb.create_sheet("RE")
    ws_re.cell(1, 1, None)
    re_headers = ["WIND", "SOLAR", "Small HYDRO", "BIO_GAS", "BIOMASS", "MSW"]
    for i, name in enumerate(re_headers, start=2):
        ws_re.cell(1, i, name)
    ws_re.cell(1, 8, None)
    ws_re.cell(1, 9, "Total")
    ws_re.cell(1, 10, None)
    # K..P headers from ENTT-2 ES..EX
    re_extra_headers = ["ES", "ET", "EU", "EV", "EW", "EX"]
    for offset, col_name in enumerate(re_extra_headers):
        entt2_idx = column_index_from_string(col_name)
        ws_re.cell(1, 11 + offset, ws_entt2.cell(1, entt2_idx).value)

    # RE header styling + alignment
    re_header_fill = PatternFill("solid", fgColor="D9D9D9")
    re_header_font = Font(bold=True, color="000000")
    re_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    re_thin = Side(style="thin", color="000000")
    re_border = Border(left=re_thin, right=re_thin, top=re_thin, bottom=re_thin)
    for c in range(1, 17):
        cell = ws_re.cell(1, c)
        if cell.value in (None, ""):
            continue
        cell.fill = re_header_fill
        cell.font = re_header_font
        cell.alignment = re_header_align
        cell.border = re_border

    # Fill RE data rows (4..99)
    re_extra_cols = [column_index_from_string(c) for c in ("ES", "ET", "EU", "EV", "EW", "EX")]
    for block in range(1, 97):
        row = block + 3
        for i, name in enumerate(re_headers, start=2):
            col_idx = entt2_col_by_name.get(name) or entt2_col_by_norm.get(norm_key(name))
            if col_idx:
                ws_re.cell(row, i, f"='ENTT-2'!{get_column_letter(col_idx)}{row}")
            else:
                ws_re.cell(row, i, 0)
        # Optional extra RE block (use blank ENTT-2 headers if any)
        for offset in range(6):
            target_col = 11 + offset
            if offset < len(re_extra_cols):
                col_idx = re_extra_cols[offset]
                ws_re.cell(row, target_col, f"='ENTT-2'!{get_column_letter(col_idx)}{row}")
            else:
                ws_re.cell(row, target_col, 0)
        ws_re.cell(row, 9, f"=SUM(B{row}:H{row})+SUM(K{row}:P{row})")
        # Align + border for data cells
        for c in range(2, 17):
            cell = ws_re.cell(row, c)
            if cell.value in (None, ""):
                continue
            cell.alignment = re_header_align
            cell.border = re_border

    # Hydel sheet (compact layout referencing ENTT-2)
    ws_h = wb.create_sheet("Hydel")
    hydel_headers = [
        None,
        "GANDHI-SAGAR HPS",
        "PENCH HPS",
        "BARGI HPS",
        " TONS HPS BSR-1",
        "BIRSINGHPUR HPS",
        "RAJGHAT HPS",
        "DEVLOND HPS BSR-3",
        "SILPARA HPS BSR-2",
        "Bargi LBCPH",
        "Jhinha BSR-4",
        "Madikheda",
        "ISP",
        "OSP",
        None,
        "Total",
    ]
    for idx, name in enumerate(hydel_headers, start=1):
        if name is not None:
            ws_h.cell(1, idx, name)

    hydel_station_names: list[str] = []
    hyd_col_by_name: dict[str, int] = {}
    for c in range(1, len(hydel_headers) + 1):
        name = ws_h.cell(1, c).value
        if name not in (None, ""):
            hyd_col_by_name[str(name)] = c
            if str(name).strip().lower() != "total":
                hydel_station_names.append(str(name))

    # Header styling for Hydel (yellow for stations + ISP/OSP)
    hyd_header_fill = PatternFill("solid", fgColor="FFFF00")
    hyd_header_font = Font(bold=True, color="000000")
    hyd_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hyd_thin = Side(style="thin", color="000000")
    hyd_border = Border(left=hyd_thin, right=hyd_thin, top=hyd_thin, bottom=hyd_thin)
    for c in range(1, len(hydel_headers) + 1):
        cell = ws_h.cell(1, c)
        if cell.value in (None, ""):
            continue
        if str(cell.value).strip().lower() != "total":
            cell.fill = hyd_header_fill
        cell.font = hyd_header_font
        cell.alignment = hyd_header_align
        cell.border = hyd_border

    for block in range(1, 97):
        row = block + 3
        for c in range(2, 15):  # B..N station/ISP/OSP
            name = ws_h.cell(1, c).value
            if name in (None, ""):
                continue
            col_idx = entt2_col_by_name.get(name) or entt2_col_by_norm.get(norm_key(name))
            if col_idx:
                ws_h.cell(row, c, f"='ENTT-2'!{get_column_letter(col_idx)}{row}")
            else:
                ws_h.cell(row, c, 0)
        # Total at P (16)
        ws_h.cell(row, 16, f"=SUM(B{row}:N{row})")

    # RE/Hydel sheets are created after ENTT-2 so they can reference its columns.

    entt1_col_by_name: dict[str, int] = {}
    if ws_entt1 is not None:
        entt1_header_row = find_timeblock_header_row(ws_entt1) or 1
        for c in range(1, ws_entt1.max_column + 1):
            name = ws_entt1.cell(entt1_header_row, c).value
            if name not in (None, ""):
                entt1_col_by_name[norm_key(name)] = c

    # Hydel ReSchedule
    ws_hr = wb.create_sheet("Hydel ReSchedule")
    ws_hr.cell(1, 1, "ISP")
    ws_hr.cell(1, 2, "OSP")
    # Column C is blank in the reference
    hydel_no_io = [n for n in hydel_station_names if n not in ("ISP", "OSP")]
    for i, name in enumerate(hydel_no_io, start=4):
        ws_hr.cell(1, i, name)
    # O/P/Q/R headers for ISP/OSP averages / reschedule inputs
    ws_hr.cell(1, 15, "ISP")
    ws_hr.cell(1, 16, "OSP")
    ws_hr.cell(1, 17, "ISP")
    ws_hr.cell(1, 18, "OSP")
    # Column S blank
    ws_hr.cell(1, 20, "Total")
    # Column U blank
    ws_hr.cell(1, 22, "Demand")
    ws_hr.cell(1, 23, "Availability")
    ws_hr.cell(1, 24, "Surplus (+) / Deficit (-)")
    # Column Y blank
    ws_hr.cell(1, 26, "Surplus (+) / Deficit (-) after reschedule Hydro")

    # Hydel ReSchedule header styling
    hr_header_fill = PatternFill("solid", fgColor="D9D9D9")
    hr_header_font = Font(bold=True, color="000000")
    hr_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hr_data_align = Alignment(horizontal="center", vertical="center")
    hr_thin = Side(style="thin", color="000000")
    hr_border = Border(left=hr_thin, right=hr_thin, top=hr_thin, bottom=hr_thin)

    hr_last_col = 31
    for c in range(1, hr_last_col + 1):
        cell = ws_hr.cell(1, c)
        if cell.value in (None, ""):
            continue
        cell.fill = hr_header_fill
        cell.font = hr_header_font
        cell.alignment = hr_header_align
        cell.border = hr_border

    # Custom header colors for ISP/OSP input columns (Q/R)
    hr_qr_header_fill = PatternFill("solid", fgColor="FFFF00")
    for c in (17, 18):
        cell = ws_hr.cell(1, c)
        if cell.value in (None, ""):
            continue
        cell.fill = hr_qr_header_fill
        cell.font = hr_header_font
        cell.alignment = hr_header_align
        cell.border = hr_border

    # Column widths for Hydel ReSchedule
    hr_widths = {
        "A": 9.1, "B": 13.0, "C": 13.0, "D": 8.9, "E": 13.0, "F": 13.0, "G": 13.0, "H": 13.0,
        "I": 13.0, "J": 13.0, "K": 13.0, "L": 13.0, "M": 13.0, "N": 8.4, "O": 9.1, "P": 13.0,
        "Q": 13.0, "R": 13.0, "S": 13.0, "T": 13.0, "U": 13.0, "V": 13.0, "W": 11.0, "X": 20.7,
        "Y": 13.0, "Z": 15.9, "AA": 13.0, "AB": 14.9, "AC": 13.0, "AD": 13.0, "AE": 13.0,
    }
    for col_letter, width in hr_widths.items():
        ws_hr.column_dimensions[col_letter].width = width

    # Precompute Hydel total column (if present)
    hyd_total_col = None
    for name, idx in hyd_col_by_name.items():
        if str(name).strip().lower() == "total":
            hyd_total_col = idx
            break

    isp_col = hyd_col_by_name.get("ISP")
    osp_col = hyd_col_by_name.get("OSP")
    if isp_col:
        isp_letter = get_column_letter(isp_col)
        ws_hr.cell(2, 15, f"=SUM(Hydel!{isp_letter}4:{isp_letter}99)/400")
        ws_hr.cell(3, 15, ArrayFormula("O3", f"=MIN(IF(Hydel!{isp_letter}4:{isp_letter}99>0, Hydel!{isp_letter}4:{isp_letter}99))"))
    else:
        ws_hr.cell(2, 15, 0)
    if osp_col:
        osp_letter = get_column_letter(osp_col)
        ws_hr.cell(2, 16, f"=SUM(Hydel!{osp_letter}4:{osp_letter}99)/400")
        ws_hr.cell(3, 16, ArrayFormula("P3", f"=MIN(IF(Hydel!{osp_letter}4:{osp_letter}99>0, Hydel!{osp_letter}4:{osp_letter}99))"))
    else:
        ws_hr.cell(2, 16, 0)
    ws_hr.cell(2, 1, "=SUM(A4:A99)/400")
    ws_hr.cell(2, 2, "=SUM(B4:B99)/400")
    ws_hr.cell(2, 17, "=SUM(O4:O99)/400")
    ws_hr.cell(2, 18, "=SUM(P4:P99)/400")
    ws_hr.cell(2, 26, 0)

    for r in (2, 3):
        for c in range(1, hr_last_col + 1):
            cell = ws_hr.cell(r, c)
            if cell.value in (None, ""):
                continue
            cell.alignment = hr_data_align
            cell.border = hr_border

    hr_q_fill = PatternFill("solid", fgColor="C9DEF1")
    hr_r_fill = PatternFill("solid", fgColor="F7C5D0")
    for block in range(1, 97):
        row = block + 3
        if isp_col:
            ws_hr.cell(row, 1, f"=Hydel!{isp_letter}{row}")
        else:
            ws_hr.cell(row, 1, 0)
        if osp_col:
            ws_hr.cell(row, 2, f"=Hydel!{osp_letter}{row}")
        else:
            ws_hr.cell(row, 2, 0)
        # copy hydel station values (from ENTT-1 when available)
        for i, name in enumerate(hydel_no_io, start=4):
            entt1_col = entt1_col_by_name.get(norm_key(name))
            if entt1_col:
                ws_hr.cell(row, i, f"='ENTT-1'!{get_column_letter(entt1_col)}{row}")
            else:
                col_idx = hyd_col_by_name.get(name)
                if col_idx:
                    ws_hr.cell(row, i, f"=Hydel!{get_column_letter(col_idx)}{row}")
                else:
                    ws_hr.cell(row, i, 0)

        # Q/R are user input columns (ISP/OSP factor). Default to 0 and color like sample.
        q_cell = ws_hr.cell(row, 17, 0)
        r_cell = ws_hr.cell(row, 18, 0)
        q_cell.fill = hr_q_fill
        r_cell.fill = hr_r_fill
        ws_hr.cell(row, 15, f"=Q{row}*$O$3")
        ws_hr.cell(row, 16, f"=R{row}*$P$3")
        ws_hr.cell(row, 20, f"=SUM(D{row}:P{row})")

        # Apply alignment + borders to Hydel ReSchedule data rows
        for c in range(1, hr_last_col + 1):
            cell = ws_hr.cell(row, c)
            if cell.value in (None, ""):
                continue
            cell.alignment = hr_data_align
            cell.border = hr_border

    # Entt sheet (match Scheduling.xlsx reference layout/formulas)
    ws_entt = wb.create_sheet("Entt")
    ws_entt.cell(1, 1, "Time Block")
    ws_entt.cell(1, 2, "from")
    ws_entt.cell(1, 3, "to")
    ws_entt.cell(1, 5, "Availability non-hydro")
    ws_entt.cell(1, 6, "Hydro")
    ws_entt.cell(1, 7, "RE")
    ws_entt.cell(1, 8, "Total")
    ws_entt.cell(3, 9, "VC")

    station_count = len(mod_stations)
    first_station_col = 10
    last_station_col = first_station_col + station_count - 1
    last_station_letter = get_column_letter(last_station_col)
    for i in range(station_count):
        col = first_station_col + i
        rank = i + 1
        col_letter = get_column_letter(col)
        ws_entt.cell(1, col, rank)
        ws_entt.cell(2, col, f"=INDEX('Name corrector'!$D:$D,MATCH({col_letter}1,'Name corrector'!$G:$G,0))")
        ws_entt.cell(3, col, f"=INDEX('Name corrector'!$F:$F,MATCH(Entt!{col_letter}1,'Name corrector'!$G:$G,0))")

    # Entt header styling (alignment + fill + borders)
    entt_header_fill = PatternFill("solid", fgColor="D9D9D9")
    entt_header_font = Font(bold=True, color="000000")
    entt_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    entt_thin = Side(style="thin", color="000000")
    entt_border = Border(left=entt_thin, right=entt_thin, top=entt_thin, bottom=entt_thin)

    for r in (1, 2, 3):
        for c in range(1, last_station_col + 1):
            cell = ws_entt.cell(r, c)
            if cell.value in (None, "") and c > 4:
                continue
            cell.fill = entt_header_fill
            cell.font = entt_header_font
            cell.alignment = entt_header_align
            cell.border = entt_border

    # Column widths for Entt sheet
    entt_widths = {
        "A": 12, "B": 10, "C": 10, "D": 2, "E": 22, "F": 10, "G": 10, "H": 10, "I": 8
    }
    for col_letter, width in entt_widths.items():
        ws_entt.column_dimensions[col_letter].width = width
    for c in range(first_station_col, last_station_col + 1):
        ws_entt.column_dimensions[get_column_letter(c)].width = 12

    # Scheduling sheet
    ws_s = wb.create_sheet("Scheduling")
    ws_s.cell(1, 1, "Time Block")
    ws_s.cell(1, 2, "from")
    ws_s.cell(1, 3, "to")
    ws_s.cell(1, 4, "Demand")
    ws_s.cell(2, 6, "Check")
    ws_s.cell(2, 7, "Hydel")
    ws_s.cell(2, 8, "RE")
    # Demand adjustment input (E3). User can change this in output to adjust demand.
    ws_s.cell(3, 5, 0)
    ws_s.cell(3, 9, "VC")
    for i, st in enumerate(stations, start=10):
        if i - 10 < len(mod_stations):
            ws_s.cell(1, i, i - 9)
        if st:
            col_letter = get_column_letter(i)
            ws_s.cell(2, i, f"=INDEX('Name corrector'!$D:$D,MATCH({col_letter}1,'Name corrector'!$G:$G,0))")
            ws_s.cell(3, i, f"=INDEX('Name corrector'!$F:$F,MATCH(Scheduling!{col_letter}1,'Name corrector'!$G:$G,0))")

    # Scheduling header styling
    sched_header_fill = PatternFill("solid", fgColor="D9D9D9")
    sched_header_font = Font(bold=True, color="000000")
    sched_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sched_thin = Side(style="thin", color="000000")
    sched_border = Border(left=sched_thin, right=sched_thin, top=sched_thin, bottom=sched_thin)

    sched_last_col = 9 + len(stations)
    for r in (1, 2, 3):
        for c in range(1, sched_last_col + 1):
            cell = ws_s.cell(r, c)
            if cell.value in (None, "") and c > 5:
                continue
            cell.fill = sched_header_fill
            cell.font = sched_header_font
            cell.alignment = sched_header_align
            cell.border = sched_border

    # Scheduling column widths
    sched_widths = {"A": 12, "B": 10, "C": 10, "D": 10, "E": 10, "F": 9, "G": 9, "H": 9, "I": 6}
    for col_letter, width in sched_widths.items():
        ws_s.column_dimensions[col_letter].width = width
    for c in range(10, sched_last_col + 1):
        ws_s.column_dimensions[get_column_letter(c)].width = 12

    # BD sheet
    ws_b = wb.create_sheet("BD")
    ws_b.cell(1, 1, "Time Block")
    ws_b.cell(1, 2, "from")
    ws_b.cell(1, 3, "to")
    ws_b.cell(1, 4, "Total BD")
    ws_b.cell(1, 5, "Availability")
    ws_b.cell(1, 6, "Demand")
    rev_stations = list(reversed(mod_stations))
    for i, st in enumerate(rev_stations, start=8):
        ws_b.cell(2, i, len(mod_stations) - (i - 8))
        col_letter = get_column_letter(i)
        ws_b.cell(3, i, f"=INDEX('Name corrector'!$D:$D,MATCH({col_letter}2,'Name corrector'!$G:$G,0))")
        ws_b.cell(4, i, f"=INDEX(Entt!3:3,MATCH(BD!{col_letter}2,Entt!1:1,0))")

    # BD header styling
    bd_header_fill = PatternFill("solid", fgColor="D9D9D9")
    bd_header_font = Font(bold=True, color="000000")
    bd_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bd_thin = Side(style="thin", color="000000")
    bd_border = Border(left=bd_thin, right=bd_thin, top=bd_thin, bottom=bd_thin)

    bd_last_col = 7 + len(mod_stations)
    for r in (1, 2, 3, 4):
        for c in range(1, bd_last_col + 1):
            cell = ws_b.cell(r, c)
            if cell.value in (None, "") and c > 6:
                continue
            cell.fill = bd_header_fill
            cell.font = bd_header_font
            cell.alignment = bd_header_align
            cell.border = bd_border

    # BD column widths
    bd_widths = {"A": 12, "B": 10, "C": 10, "D": 10, "E": 13, "F": 10, "G": 2}
    for col_letter, width in bd_widths.items():
        ws_b.column_dimensions[col_letter].width = width
    for c in range(8, bd_last_col + 1):
        ws_b.column_dimensions[get_column_letter(c)].width = 12

    # IEX and Bid Sheet layouts (need before Deamnd-MP loop)
    ws_bids = wb.create_sheet("Bid Sheet")
    ws_iex = wb.create_sheet("IEX")

    # Bid sheet header/layout (align to reference)
    date_val = None
    for sheet_name in ("Demand-EZ", "Demand-CZ", "Demand-WZ"):
        if sheet_name in wb.sheetnames:
            date_val = wb[sheet_name].cell(2, 5).value
            if date_val:
                break
    setup_bid_sheet_layout(
        ws_bids,
        date_val,
        BID_PRICE_POINTS,
        BID_MARGIN,
        "Demand-EZ" in wb.sheetnames,
    )

    # IEX header
    ws_iex.cell(3, 2, "Period")
    ws_iex.cell(3, 3, "Qty in MW")
    ws_iex.cell(3, 4, "Rate/MWh")
    ws_iex.cell(3, 5, "Amount in `")
    iex_header_fill = PatternFill("solid", fgColor="FF8080FF")
    iex_header_font = Font(bold=True, color="000000")
    iex_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    iex_thin = Side(style="thin", color="000000")
    iex_border = Border(left=iex_thin, right=iex_thin, top=iex_thin, bottom=iex_thin)
    for c in range(2, 6):
        cell = ws_iex.cell(3, c)
        cell.fill = iex_header_fill
        cell.font = iex_header_font
        cell.alignment = iex_header_align
        cell.border = iex_border

    # Deamnd-MP sheet
    ws_dm = wb.create_sheet("Deamnd-MP")
    # Row 1-2 intentionally blank to match reference layout
    ws_dm.cell(3, 1, "Block.      No.")
    ws_dm.cell(3, 2, "Time block")
    ws_dm.cell(3, 4, "EZ")
    ws_dm.cell(3, 5, "CZ")
    ws_dm.cell(3, 6, "WZ")
    ws_dm.cell(3, 7, "Total")
    ws_dm.cell(4, 2, "From")
    ws_dm.cell(4, 3, "To")

    # Styling for Deamnd-MP sheet (alignment, borders, header colors)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True, color="000000")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, width in {"A": 13, "B": 13, "C": 13, "D": 9.1, "E": 13, "F": 13, "G": 13}.items():
        ws_dm.column_dimensions[col].width = width

    for c in range(1, 8):
        cell = ws_dm.cell(3, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border_all
        cell = ws_dm.cell(4, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border_all

    # Column maps for scheduling/BD formulas
    entt_col_by_station = {st: first_station_col + idx for idx, st in enumerate(mod_stations)}
    sched_col_by_station = {st: first_station_col + idx for idx, st in enumerate(stations) if st}
    last_sched_col = 9 + len(stations)
    last_sched_letter = get_column_letter(last_sched_col)
    last_bd_col = 7 + len(mod_stations)
    last_bd_letter = get_column_letter(last_bd_col)
    bid_start_col = 3
    bid_last_col = bid_start_col + len(BID_PRICE_POINTS) - 1
    bid_last_letter = get_column_letter(bid_last_col)

    for block in range(1, 97):
        row_dm = block + 4
        row_std = block + 3
        d_cz = demand_cz.get(block, {})
        d_ez = demand_ez.get(block, {})
        d_wz = demand_wz.get(block, {})
        from_t = d_cz.get("from") or d_ez.get("from") or d_wz.get("from")
        to_t = d_cz.get("to") or d_ez.get("to") or d_wz.get("to")
        demand = safe_float(d_cz.get("demand")) + safe_float(d_ez.get("demand")) + safe_float(d_wz.get("demand"))

        ws_dm.cell(row_dm, 1, block)
        ws_dm.cell(row_dm, 2, from_t)
        ws_dm.cell(row_dm, 3, to_t)
        ws_dm.cell(row_dm, 4, safe_float(d_ez.get("demand")))
        ws_dm.cell(row_dm, 5, safe_float(d_cz.get("demand")))
        ws_dm.cell(row_dm, 6, safe_float(d_wz.get("demand")))
        ws_dm.cell(row_dm, 7, demand)

        for c in range(1, 8):
            cell = ws_dm.cell(row_dm, c)
            cell.alignment = data_align
            cell.border = border_all
        ws_dm.cell(row_dm, 2).number_format = "h:mm"
        ws_dm.cell(row_dm, 3).number_format = "h:mm"

        # Bid Sheet + IEX periods
        def fmt_time(t: object) -> str:
            if hasattr(t, "strftime"):
                return t.strftime("%H:%M")
            s = str(t or "").strip()
            if s:
                return s[:5]
            return ""

        period = f"{fmt_time(from_t)} - {fmt_time(to_t)}" if from_t and to_t else ""
        row_bid = block + 5
        ws_bids.cell(row_bid, 1, block)
        ws_bids.cell(row_bid, 2, period)

        ws_iex.cell(row_std, 2, period)
        iex_row = iex_data.get(period, {})
        ws_iex.cell(row_std, 3, iex_row.get("qty", 0.0))
        ws_iex.cell(row_std, 4, iex_row.get("rate", 0.0))
        ws_iex.cell(row_std, 5, iex_row.get("amount", 0.0))
        for c in (3, 4, 5):
            ws_iex.cell(row_std, c).number_format = "0.00"

        # Entt rows
        ws_entt.cell(row_std, 1, block)
        ws_entt.cell(row_std, 2, from_t)
        ws_entt.cell(row_std, 3, to_t)
        ws_entt.cell(row_std, 5, f"=SUM(J{row_std}:{last_station_letter}{row_std})")
        ws_entt.cell(row_std, 6, f"=Scheduling!G{row_std}")
        ws_entt.cell(row_std, 7, f"=Scheduling!H{row_std}")
        ws_entt.cell(row_std, 8, f"=E{row_std}+F{row_std}+G{row_std}")
        for i in range(station_count):
            col = first_station_col + i
            col_letter = get_column_letter(col)
            ws_entt.cell(
                row_std,
                col,
                f"=INDEX('ENTT-2'!{row_std}:{row_std},MATCH(Entt!{col_letter}$2,'ENTT-2'!$1:$1,0))",
            )
        # Scheduling (formulas so E3 demand adjustment flows through)
        ws_s.cell(row_std, 1, block)
        ws_s.cell(row_std, 2, from_t)
        ws_s.cell(row_std, 3, to_t)
        ws_s.cell(row_std, 4, demand)
        # Column E should update if user changes E3 in Excel
        ws_s.cell(row_std, 5, f"=D{row_std}+$E$3")

        for i, st in enumerate(stations, start=10):
            if not st:
                continue
            ent_val = safe_float(ent.get(block, {}).get(st, 0.0))
            col_letter = get_column_letter(i)
            if i == first_station_col:
                remaining_expr = f"$E{row_std}-$G{row_std}-$H{row_std}"
            else:
                prev_letter = get_column_letter(i - 1)
                remaining_expr = f"$E{row_std}-$G{row_std}-$H{row_std}-SUM(${get_column_letter(first_station_col)}{row_std}:{prev_letter}{row_std})"
            ws_s.cell(row_std, i, f"=MAX(MIN({ent_val:.6f},{remaining_expr}),0)")

        ws_s.cell(row_std, 6, f"=SUM(J{row_std}:{last_sched_letter}{row_std})")
        ws_s.cell(row_std, 7, f"='Hydel ReSchedule'!T{row_std}")
        ws_s.cell(row_std, 8, f"=RE!I{row_std}")
        ws_s.cell(row_std, 9, f"=IF(ABS((F{row_std}+G{row_std}+H{row_std})-E{row_std})<0.000001,\"OK\",\"Error\")")

        # BD (data rows start at row 5 to keep VC row in row 4)
        row_b = row_std + 1
        ws_b.cell(row_b, 1, block)
        ws_b.cell(row_b, 2, from_t)
        ws_b.cell(row_b, 3, to_t)
        # Availability should mirror Entt sheet Total column
        ws_b.cell(row_b, 5, f"=Entt!H{row_std}")
        ws_b.cell(row_b, 6, f"=Scheduling!E{row_std}")
        for i, st in enumerate(rev_stations, start=8):
            entt_col = entt_col_by_station.get(st)
            sched_col = sched_col_by_station.get(st)
            if entt_col and sched_col:
                entt_letter = get_column_letter(entt_col)
                sched_letter = get_column_letter(sched_col)
                ws_b.cell(row_b, i, f"=Entt!{entt_letter}{row_std}-Scheduling!{sched_letter}{row_std}")
            else:
                ws_b.cell(row_b, i, 0)
        ws_b.cell(row_b, 4, f"=SUM(H{row_b}:{last_bd_letter}{row_b})")

        # Bid Sheet formulas (match reference SUMIF logic)
        criteria_range = f"BD!$H$4:${last_bd_letter}$4"
        sum_range = f"BD!$H{row_b}:${last_bd_letter}{row_b}"
        for c in range(bid_start_col, bid_last_col + 1):
            col_letter = get_column_letter(c)
            formula = (
                f"=IF((SUMIF({criteria_range},\"<\"&{col_letter}$5-$O$3,{sum_range}))=0,"
                f"\"\",(SUMIF({criteria_range},\"<\"&{col_letter}$5-$O$3,{sum_range})))"
            )
            cell = ws_bids.cell(row_bid, c, formula)
            cell.number_format = "0"

        # Hydel ReSchedule computed fields
        ws_hr.cell(row_std, 22, f"=Scheduling!E{row_std}")
        ws_hr.cell(row_std, 23, f"=Entt!E{row_std}+Entt!G{row_std}+Entt!F{row_std}")
        ws_hr.cell(row_std, 24, f"=W{row_std}-V{row_std}")
        ws_hr.cell(row_std, 26, f"=W{row_std}-V{row_std}-Entt!F{row_std}+'Hydel ReSchedule'!T{row_std}")
        # Extra trailing columns (AD/AE) hold time window
        ws_hr.cell(row_std, 30, from_t)
        ws_hr.cell(row_std, 31, to_t)
        for c in (22, 23, 24, 26, 30, 31):
            cell = ws_hr.cell(row_std, c)
            cell.alignment = hr_data_align
            cell.border = hr_border
            if c in (30, 31):
                cell.number_format = "h:mm"

    # Bid sheet formulas already populated during main loop

    # Hydel ReSchedule period averages and labels (rows 4-7)
    period_slices = [
        (4, "00.00 to 06.00", 4, 27),
        (5, "06.00 to 11.00", 28, 47),
        (6, "11.00 to 18.00", 48, 75),
        (7, "18.00 to 24.00", 76, 99),
    ]
    for row, label, start_row, end_row in period_slices:
        ws_hr.cell(row, 27, f"=AVERAGE(X{start_row}:X{end_row})")
        ws_hr.cell(row, 28, label)
        for c in (27, 28):
            cell = ws_hr.cell(row, c)
            cell.alignment = hr_data_align
            cell.border = hr_border

    # Check sheet (simple OK markers)
    ws_chk = wb.create_sheet("Check")
    ws_chk.cell(1, 1, "Time Block")
    ws_chk.cell(1, 2, "from")
    ws_chk.cell(1, 3, "to")
    ws_chk.cell(1, 5, "Availability non-hydro")
    for i, st in enumerate(mod_stations, start=8):
        ws_chk.cell(1, i, i - 7)
        ws_chk.cell(2, i, st)
        ws_chk.cell(3, i, vc_by_station.get(st, 0.0))
    ws_chk.cell(3, 7, "VC")

    # add 4 trailing blank columns to match reference width
    last_col = 8 + len(mod_stations) - 1
    for extra in range(1, 5):
        ws_chk.cell(1, last_col + extra, "")
    for block in range(1, 97):
        row = block + 3
        ws_chk.cell(row, 1, block)
        ws_chk.cell(row, 2, ws_s.cell(row, 2).value)
        ws_chk.cell(row, 3, ws_s.cell(row, 3).value)
        ws_chk.cell(row, 5, ws_entt.cell(row, 5).value)
        for i, _ in enumerate(mod_stations, start=8):
            ws_chk.cell(row, i, "OK")
        for extra in range(1, 5):
            ws_chk.cell(row, last_col + extra, "")

    # IEX/Bid sheets already created above

    # Extend Deamnd-MP to match reference row count
    ws_dm.cell(101, 1, "")
    ws_dm.cell(102, 1, "")

    # PowerBI data sheet (summary tables for charts/KPIs)
    ws_pbi = wb.create_sheet("PowerBI_Data")
    pbi_headers = [
        "Time Block",
        "From",
        "To",
        "Demand",
        "Scheduled Supply",
        "Surplus Before",
        "Surplus After",
        "Hydro",
        "Renewables",
        "Gas",
        "Thermal",
    ]
    for col, header in enumerate(pbi_headers, start=1):
        ws_pbi.cell(1, col, header)

    # Identify gas stations by code suffix
    gas_suffix_re = re.compile(r"_(APM|NAPM|RLNG|CRF|GF|RF|LF)$", re.IGNORECASE)
    gas_stations = [st for st in stations if st and gas_suffix_re.search(st)]

    # Gas station list (for SUMPRODUCT)
    gas_list_col = 13  # Column M
    ws_pbi.cell(1, gas_list_col, "Gas Stations")
    for idx, name in enumerate(gas_stations, start=2):
        ws_pbi.cell(idx, gas_list_col, name)

    gas_list_start = 2
    gas_list_end = max(gas_list_start, gas_list_start + len(gas_stations) - 1)
    gas_list_range = f"${get_column_letter(gas_list_col)}${gas_list_start}:${get_column_letter(gas_list_col)}${gas_list_end}"

    for block in range(1, 97):
        row_pbi = block + 1
        row_std = block + 3
        ws_pbi.cell(row_pbi, 1, f"=Scheduling!A{row_std}")
        ws_pbi.cell(row_pbi, 2, f"=Scheduling!B{row_std}")
        ws_pbi.cell(row_pbi, 3, f"=Scheduling!C{row_std}")
        # Use adjusted demand so changes in Scheduling!E3 flow into PowerBI
        ws_pbi.cell(row_pbi, 4, f"=Scheduling!E{row_std}")
        ws_pbi.cell(row_pbi, 5, f"=Scheduling!F{row_std}+Scheduling!G{row_std}+Scheduling!H{row_std}")
        ws_pbi.cell(row_pbi, 6, f"='Hydel ReSchedule'!X{row_std}")
        ws_pbi.cell(row_pbi, 7, f"='Hydel ReSchedule'!Z{row_std}")
        ws_pbi.cell(row_pbi, 8, f"='Hydel ReSchedule'!T{row_std}")
        ws_pbi.cell(row_pbi, 9, f"=RE!I{row_std}")

        gas_formula = (
            f"=SUMPRODUCT(--ISNUMBER(MATCH(Scheduling!$J$2:${last_sched_letter}$2,{gas_list_range},0))"
            f",Scheduling!$J{row_std}:${last_sched_letter}{row_std})"
        )
        ws_pbi.cell(row_pbi, 10, gas_formula)
        ws_pbi.cell(row_pbi, 11, f"=Scheduling!F{row_std}-J{row_pbi}")

    # Generation mix table (for pie chart)
    mix_start = 102
    ws_pbi.cell(mix_start, 1, "Generation Mix (Total MW)")
    ws_pbi.cell(mix_start + 1, 1, "Plant Type")
    ws_pbi.cell(mix_start + 1, 2, "MW")
    mix_rows = [
        ("Thermal", f"=SUM(K2:K97)"),
        ("Hydro", f"=SUM(H2:H97)"),
        ("Gas", f"=SUM(J2:J97)"),
        ("Renewables", f"=SUM(I2:I97)"),
    ]
    for idx, (label, formula) in enumerate(mix_rows, start=0):
        ws_pbi.cell(mix_start + 2 + idx, 1, label)
        ws_pbi.cell(mix_start + 2 + idx, 2, formula)

    # KPI table
    kpi_start = 110
    ws_pbi.cell(kpi_start, 1, "KPI")
    ws_pbi.cell(kpi_start + 1, 1, "Peak Demand")
    ws_pbi.cell(kpi_start + 1, 2, "=MAX(D2:D97)")
    ws_pbi.cell(kpi_start + 2, 1, "MOD Rate (Rs/MWh)")
    ws_pbi.cell(kpi_start + 2, 2, "=AVERAGEIF('Name corrector'!F:F,\">0\")")
    ws_pbi.cell(kpi_start + 3, 1, "Cost Savings")
    ws_pbi.cell(
        kpi_start + 3,
        2,
        "=(0-SUMIF(F2:F97,\"<0\")-(0-SUMIF(G2:G97,\"<0\")))*B112",
    )
    ws_pbi.cell(kpi_start + 4, 1, "Market Price (Rs/MWh)")
    ws_pbi.cell(kpi_start + 4, 2, "=AVERAGE(IEX!D4:D99)")
    ws_pbi.cell(kpi_start + 5, 1, "Market Profit")
    ws_pbi.cell(kpi_start + 5, 2, "=SUMIF(G2:G97,\">0\")*B114")

    # Basic formatting for PowerBI sheet
    pbi_header_fill = PatternFill("solid", fgColor="D9D9D9")
    pbi_header_font = Font(bold=True, color="000000")
    pbi_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    pbi_thin = Side(style="thin", color="000000")
    pbi_border = Border(left=pbi_thin, right=pbi_thin, top=pbi_thin, bottom=pbi_thin)
    for c in range(1, 12):
        cell = ws_pbi.cell(1, c)
        cell.fill = pbi_header_fill
        cell.font = pbi_header_font
        cell.alignment = pbi_header_align
        cell.border = pbi_border
    for r in range(2, 98):
        for c in range(1, 12):
            cell = ws_pbi.cell(r, c)
            cell.alignment = pbi_header_align
            cell.border = pbi_border
    for col_letter, width in {
        "A": 12, "B": 10, "C": 10, "D": 12, "E": 16, "F": 14,
        "G": 14, "H": 10, "I": 12, "J": 10, "K": 12, "M": 20,
    }.items():
        ws_pbi.column_dimensions[col_letter].width = width

    # Auto-fit header widths for all sheets (only if not explicitly set)
    autofit_header_widths(wb)

    # Remove sheets that are not linked to the final logic/output
    prune_unlinked_sheets(wb)
    reorder_sheets(
        wb,
        [
            "Scheduling",
            "BD",
            "Hydel ReSchedule",
            "Bid Sheet",
        ],
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Scheduling + BD Generator")
        self.geometry("920x440")
        self.resizable(False, False)

        cwd = Path.cwd()
        self.paths = {
            "Demand-CZ.xlsx": tk.StringVar(value=""),
            "Demand-EZ.xlsx": tk.StringVar(value=""),
            "Demand-WZ.xlsx": tk.StringVar(value=""),
            "IEX.pdf": tk.StringVar(value=""),
            "Entitlement.xlsx": tk.StringVar(value=""),
            "MoD_Rate.xlsx": tk.StringVar(value=""),
            "Output": tk.StringVar(value=str(cwd / "Final_Scheduling_BD_Output.xlsx")),
        }

        self._build_ui()

    def _pick_file(self, key: str):
        if key.lower().endswith(".pdf"):
            types = [("PDF files", "*.pdf"), ("All files", "*.*")]
        else:
            types = [("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        path = filedialog.askopenfilename(
            title=f"Select {key}",
            filetypes=types,
        )
        if path:
            self.paths[key].set(path)

    def _pick_output(self):
        path = filedialog.asksaveasfilename(
            title="Save Output Workbook",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="Final_Scheduling_BD_Output.xlsx",
        )
        if path:
            self.paths["Output"].set(path)

    def _build_ui(self):
        padx = 12
        tk.Label(self, text="Input Files", font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=padx, pady=(10, 4))

        for key in ["Demand-CZ.xlsx", "Demand-EZ.xlsx", "Demand-WZ.xlsx", "IEX.pdf", "Entitlement.xlsx", "MoD_Rate.xlsx"]:
            row = tk.Frame(self)
            row.pack(fill="x", padx=padx, pady=4)
            tk.Label(row, text=key, width=18, anchor="w").pack(side="left")
            tk.Entry(row, textvariable=self.paths[key]).pack(side="left", fill="x", expand=True)
            tk.Button(row, text="Browse", command=lambda k=key: self._pick_file(k), width=10).pack(side="left", padx=(8, 0))

        tk.Label(self, text="Output", font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=padx, pady=(16, 4))
        out_row = tk.Frame(self)
        out_row.pack(fill="x", padx=padx, pady=4)
        tk.Label(out_row, text="Output file", width=18, anchor="w").pack(side="left")
        tk.Entry(out_row, textvariable=self.paths["Output"]).pack(side="left", fill="x", expand=True)
        tk.Button(out_row, text="Browse", command=self._pick_output, width=10).pack(side="left", padx=(8, 0))
        tk.Button(self, text="Generate Scheduling + BD", command=self.run, height=2).pack(pady=20)
        self.status = tk.Label(self, text="Ready", fg="#1d4ed8", anchor="w")
        self.status.pack(fill="x", padx=padx)

    def run(self):
        try:
            self.status.config(text="Processing...", fg="#b45309")
            self.update_idletasks()

            demand_paths = {
                "Demand-CZ.xlsx": Path(self.paths["Demand-CZ.xlsx"].get().strip()) if self.paths["Demand-CZ.xlsx"].get().strip() else None,
                "Demand-EZ.xlsx": Path(self.paths["Demand-EZ.xlsx"].get().strip()) if self.paths["Demand-EZ.xlsx"].get().strip() else None,
                "Demand-WZ.xlsx": Path(self.paths["Demand-WZ.xlsx"].get().strip()) if self.paths["Demand-WZ.xlsx"].get().strip() else None,
            }
            provided_demands = [p for p in demand_paths.values() if p is not None]
            if len(provided_demands) < 2:
                raise ValueError("Please provide at least two Demand files (out of CZ, EZ, WZ).")

            build_outputs(
                demand_paths["Demand-CZ.xlsx"],
                demand_paths["Demand-EZ.xlsx"],
                demand_paths["Demand-WZ.xlsx"],
                Path(self.paths["IEX.pdf"].get().strip()) if self.paths["IEX.pdf"].get().strip() else None,
                Path(self.paths["Entitlement.xlsx"].get().strip()),
                Path(self.paths["MoD_Rate.xlsx"].get().strip()),
                Path(self.paths["Output"].get().strip()),
            )

            out = self.paths["Output"].get().strip()
            self.status.config(text=f"Done. Output created: {out}", fg="#15803d")
            messagebox.showinfo("Success", f"Output created:\n{out}")
        except Exception as exc:
            self.status.config(text="Failed", fg="#b91c1c")
            messagebox.showerror("Error", str(exc))


if __name__ == "__main__":
    App().mainloop()
